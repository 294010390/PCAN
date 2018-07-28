# 读写2007 excel
import openpyxl
import os
import json

f = open("Excel_Config.json", encoding='utf-8')  #设置以utf-8解码模式读取文件，encoding参数必须设置，否则默认以gbk模式读取文件，当文件中包含中文时，会报错
Config = json.load(f)

StartPoint = Config['StartPoint']   #注意多重结构的读取语法
Message = Config['Message']
CAN_ID = Config['CAN_ID']
Type = Config['Type']
DiagConnection = Config['DiagConnection']
Signal = Config['Signal']
Short_Name = Config['Short_Name']
Start_Byte = Config['Start_Byte']
Start_Bit = Config['Start_Bit']
Len = Config['Len']
Data = Config['Data']
Range = Config['Range']
Conversion = Config['Conversion']
DLC = Config['DLC']
CycleMessage = Config['CycleMessage']
CycleCAN_ID = Config['CycleCAN_ID']
CyclePeriodic = Config['CyclePeriodic']

Message_Table = {}
AllValue = []

def readIB_MsgSig(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name('IB_MsgSig')
    # 每一列的标志

    CAN_ID_Name = "CAN_ID"
    Sig_Short_Name = "Sig_Short_Name"
    Sig_Value_Name = "Sig_Value_Name"

    global Message_Table
    global AllValue

    SigValueDict = {}
    RangeValueDict = {}
    LocationandDLCDic = {}
    SigNameList = []
    SigValueList = []

    PreMessageName = None
    PreCAN_IDName = None

    for i in range(StartPoint, sheet.max_row):

        Value = []
        AllValue = []

        MessageName = sheet[Message + str(i)].value
        CAN_ID_Value = sheet[CAN_ID + str(i)].value
        Sig_ShortName_Value = sheet[Short_Name + str(i)].value
        DataValue = sheet[Data + str(i)].value
        RangeValue = sheet[Range + str(i)].value
        SigValue = sheet[Conversion + str(i)].value
        TypeValue = sheet[Type + str(i)].value
        StartByteValue = sheet[Start_Byte + str(i)].value
        StartBitValue = sheet[Start_Bit + str(i)].value
        LenValue = sheet[Len + str(i)].value
        DLCValue = sheet[DLC + str(i)].value

        if PreMessageName == None:
            PreMessageName = MessageName
            PreCAN_IDName = CAN_ID_Value

        if PreMessageName == MessageName or MessageName == None:
            # MessageName = PreMessageName
            SigNameList.append(Sig_ShortName_Value)
            # SigValueList.append(SigValueDict)
        else:
            PreMessageName = MessageName
            PreCAN_IDName = CAN_ID_Value
            SigNameList = []
            # SigValueList = []
            SigValueDict = {}
            RangeValueDict = {}
            LocationandDLCDic = {}
            SigNameList.append(Sig_ShortName_Value)

        LocationandDLCDic[Sig_ShortName_Value] = (StartByteValue, StartBitValue, LenValue, DLCValue)

        if DataValue == 'BLN':
            Value = [("False", 0), ("True", 1)]
            SigValueDict[Sig_ShortName_Value] = Value
            RangeValueDict[Sig_ShortName_Value] = ['NA']
        elif DataValue == 'ENM':
            newSigValue = SigValue.split(";")
            for i in newSigValue:
                j = i.split('=')
                k = j[0][-1]
                v = j[1]
                Value.append((k, v))
            SigValueDict[Sig_ShortName_Value] = Value
            RangeValueDict[Sig_ShortName_Value] = ['NA']
        elif DataValue == 'SNM' or DataValue == 'UNM':
            Value = ['NA']
            SigValueDict[Sig_ShortName_Value] = Value
            nRangeValueList = RangeValue.split(' ')
            RangeValueDict[Sig_ShortName_Value] = (nRangeValueList[0], nRangeValueList[2])

        elif DataValue == 'PKT' or DataValue == 'ASC':
            Value = ['NA']
            SigValueDict[Sig_ShortName_Value] = Value
            RangeValueDict[Sig_ShortName_Value] = Value

        # elif DataValue == 'ASC':
        #     Value = ['NA']
        #     SigValueDict[Sig_ShortName_Value] = Value
        #     RangeValueList = Value

        AllValue.append(PreCAN_IDName)
        AllValue.append(TypeValue)
        AllValue.append(SigNameList)
        AllValue.append(RangeValueDict)
        AllValue.append(SigValueDict)
        AllValue.append(LocationandDLCDic)

        Message_Table[PreMessageName] = AllValue

def readCycleTime(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name('IB_Tx')

    global Message_Table
    global AllValue

    for i in range(StartPoint, sheet.max_row):
        AllValue = []

        CycleMessageValue = sheet[CycleMessage + str(i)].value
        CycleCAN_IDValue = sheet[CycleCAN_ID + str(i)].value
        CyclePeriodicValue = sheet[CyclePeriodic + str(i)].value

        if CyclePeriodicValue == '0':
            CyclePeriodicValue = '10.0'

        if CycleMessageValue != None and Message_Table[CycleMessageValue][-1] != CyclePeriodicValue:
            # Message_Table[CycleMessageValue].append(CycleCAN_IDValue)
            Message_Table[CycleMessageValue].append(CyclePeriodicValue)

            # print('My AllValue is ', Message_Table[CycleMessageValue])

        # Message_Table[CycleMessageValue] = AllValue
    # for k, v in Message_Table.items():
    #     print(k, "---", v)

def ReturnPath():
    filepath = ''
    dir = os.getcwd()
    for name in os.listdir(dir):
        if "xlsx" in name:
            filepath = os.path.join(dir, name)
            break
    return filepath

def ReturnMessage_Table():
    filepath = ReturnPath()
    readIB_MsgSig(filepath)
    readCycleTime(filepath)
    return Message_Table

if __name__ == '__main__' :
    # filepath = ReturnPath()
    # readIB_MsgSig(filepath)
    # readCycleTime(filepath)
    table = ReturnMessage_Table()
    s = 0
    for i,j in table.items():
        s += len(j[2])
    print(s)

    print(len(table))
    for k, v in table.items():
        print(k, "---", v)

        # 检查获取的数据是否为空
        # if k == None:
        #     print(k, "---", v)
        # if None in v:
        #     print(k, "---", v)
